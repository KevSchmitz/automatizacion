from openpyxl import load_workbook
import tempfile
from constants import BASE_DIR

def fill_template(data):
    template_path = BASE_DIR / 'template.xlsx'

    # 1.Crear e inyectar datos en la plantilla
    wb = load_workbook(template_path)
    ws = wb['Plantilla Desarrollo']

    id_cell = ws["F2"]
    producto_cell = ws["D4"]
    cliente_cell = ws['C6']
    solicitante_cell = ws["H6"]

    for row in data:
        id_cell.value = row['n desarrollo']
        producto_cell.value = row ['producto']
        cliente_cell.value = row["cliente"]
        solicitante_cell.value = row["solicitante"]

    temp_file_path = ""

    # 2. Guardar el archivo temporal con los datos inyectados
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        temp_file_path = tmp.name
        wb.save(temp_file_path) # Se guarda temporalmente el archivo

    return temp_file_path


   
